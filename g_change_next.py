import streamlit as st
import pandas as pd
import re
import unicodedata
import io
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# ===============================
# 簡易ログイン（パスワード認証）
# ===============================
def check_password():
    """st.secrets['password'] と一致するかを確認する簡易ログイン"""
    def password_entered():
        """テキストボックスに入力されたパスワードを検証"""
        if "password" not in st.secrets:
            st.session_state["password_correct"] = False
            st.session_state["password_error"] = "サーバー側にパスワードが設定されていません。管理者に確認してください。"
            return

        if st.session_state["password"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            st.session_state.pop("password", None)  # パスワード文字列は消しておく
            st.session_state.pop("password_error", None)
        else:
            st.session_state["password_correct"] = False
            st.session_state["password_error"] = "パスワードが違います。もう一度入力してください。"

    # 初回：パスワード入力欄を表示
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False

    if not st.session_state["password_correct"]:
        st.title("🔐 G-Change Next ログイン")
        st.text_input(
            "パスワードを入力してください",
            type="password",
            on_change=password_entered,
            key="password",
        )
        if "password_error" in st.session_state and st.session_state["password_error"]:
            st.error(st.session_state["password_error"])
        # ここで処理をストップ（アプリ本体はまだ表示しない）
        return False

    # 認証済み
    return True


# ===============================
# Streamlit設定
# ===============================
st.set_page_config(page_title="G-Change Next", layout="wide")

# ▼ここでログインチェック。失敗したら以降の処理は実行されない
if not check_password():
    st.stop()

st.title("🚗 G-Change Next｜企業情報整形＆NG除外ツール（Ver6.3 複数ファイル対応＋確定ボタン省略版）")

# ===============================
# テキスト正規化
# ===============================
def nfkc(s: str) -> str:
    return unicodedata.normalize("NFKC", s)

def normalize_text(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).replace("\u3000", " ").replace("\xa0", " ")
    s = re.sub(r'[−–—―ー]', '-', s)
    return nfkc(s).strip()

def clean_address(address: str) -> str:
    address = normalize_text(address)
    return address.strip()

def extract_industry(line: str) -> str:
    return normalize_text(line)

# ===============================
# 企業名正規化（NG照合用）
# ===============================
COMPANY_SUFFIXES = ["株式会社", "(株)", "（株）", "有限会社", "(有)", "（有）", "合同会社"]
def canonical_company_name(name: str) -> str:
    s = normalize_text(name)
    for suf in sorted(COMPANY_SUFFIXES, key=len, reverse=True):
        s = s.replace(suf, "")
    s = re.sub(r"[\s\-・/,.·･\(\)（）【】＆&＋+_|]", "", s)
    return s

# ===============================
# 電話番号処理（原文保持）
# ===============================
HYPHENS = "-‒–—―−－ー‐﹣\u2011"
HYPHENS_CLASS = re.escape(HYPHENS)

# 電話番号候補抽出（誤検出防止）: 数字＋ハイフン/空白が続く8文字以上の塊
CANDIDATE_RE = re.compile(rf"[+]?\d(?:[\d{HYPHENS_CLASS}\s]{{6,}})\d")

def pick_phone_token_raw(line: str) -> str:
    """1行から電話番号らしい文字列を抽出。digits 長が 9〜11 以外は不採用。原文表記（ハイフン位置）をそのまま返す。"""
    if not line:
        return ""
    s = unicodedata.normalize("NFKC", str(line))
    raw_cands = CANDIDATE_RE.findall(s)
    cands = []
    for token in raw_cands:
        tok = token.strip()
        if ":" in tok:           # 時刻混入などは除外
            continue
        digits = re.sub(r"\D", "", tok)
        if not (9 <= len(digits) <= 11):
            continue             # 11-10 のような短い塊は除外
        if not (digits.startswith("0") or digits.startswith("81")):
            continue             # 国内先頭0 or 国番号81のみ許可
        score = (len(digits), tok.count("-"))  # 長いdigits＆ハイフン多い＝電話っぽい
        cands.append((score, tok))
    if not cands:
        return ""
    cands.sort(key=lambda x: x[0], reverse=True)
    return cands[0][1]

def phone_digits_only(s: str) -> str:
    """内部照合用に数字だけ抽出（原文表記は保持）"""
    return re.sub(r"\D", "", str(s or ""))

# ===============================
# 抽出プロファイル（既存3方式）
# ===============================
# 1) Google検索リスト（縦読み・電話上下）
def extract_google_vertical(lines):
    results = []
    rows = [str(l) for l in lines if str(l).strip() != ""]
    for i, line in enumerate(rows):
        ph_raw = pick_phone_token_raw(line)
        if ph_raw:
            phone = ph_raw  # 原文保持
            address = rows[i - 1] if i - 1 >= 0 else ""
            industry = extract_industry(rows[i - 2]) if i - 2 >= 0 else ""
            company = rows[i - 3] if i - 3 >= 0 else ""
            results.append([company, industry, clean_address(address), phone])
    return pd.DataFrame(results, columns=["企業名", "業種", "住所", "電話番号"])

# 2) シゴトアルワ（縦積み）
def extract_shigoto_arua(df_like: pd.DataFrame) -> pd.DataFrame:
    df = df_like.copy()
    if df.columns.size > 2:
        df = df.iloc[:, :2]
    df.columns = ["col0", "col1"]
    df = df.fillna("")
    current = {"企業名": "", "住所": "", "電話番号": "", "業種": ""}
    out = []

    def flush():
        if current["企業名"]:
            out.append([current["企業名"], current["業種"], current["住所"], current["電話番号"]])
        current.update({"企業名": "", "住所": "", "電話番号": "", "業種": ""})

    for _, row in df.iterrows():
        k, v = str(row["col0"]), str(row["col1"])
        if k in ["住所", "所在地", "本社所在地"]:
            current["住所"] = clean_address(v)
        elif k in ["電話", "電話番号", "TEL", "Tel", "tel"]:
            current["電話番号"] = v  # 原文保持
        elif k in ["業種", "事業内容", "産業分類", "製造業種"]:
            current["業種"] = extract_industry(v)
        elif k and not v:
            if current["企業名"]:
                flush()
            current["企業名"] = k
    if current["企業名"]:
        flush()
    return pd.DataFrame(out, columns=["企業名", "業種", "住所", "電話番号"])

# 3) 日本倉庫協会（4列）
def extract_warehouse_association(df_like: pd.DataFrame) -> pd.DataFrame:
    df = df_like.fillna("")
    if df.shape[1] < 2:
        return pd.DataFrame(columns=["企業名", "業種", "住所", "電話番号"])
    while df.shape[1] < 4:
        df[f"__pad{df.shape[1]}"] = ""
    df = df.iloc[:, :4]
    df.columns = ["c0", "c1", "c2", "c3"]

    tel_re = re.compile(r"(?:TEL|Tel|tel)\s*([0-9０-９\-ー－\s]+)")
    out, current = [], {"企業名": "", "住所": "", "電話番号": "", "業種_set": set()}

    def flush():
        if current["企業名"]:
            out.append([current["企業名"], "・".join(current["業種_set"]), current["住所"], current["電話番号"]])
        current.update({"企業名": "", "住所": "", "電話番号": "", "業種_set": set()})

    for _, r in df.iterrows():
        if r["c0"]:
            if current["企業名"] and r["c0"] != current["企業名"]:
                flush()
            current["企業名"] = r["c0"]
        if r["c1"]:
            current["住所"] = clean_address(r["c1"])
        if r["c2"]:
            m = tel_re.search(r["c2"])
            if m:
                current["電話番号"] = m.group(1).strip()  # 原文保持
        if r["c3"]:
            current["業種_set"].add(extract_industry(r["c3"]))
    if current["企業名"]:
        flush()
    return pd.DataFrame(out, columns=["企業名", "業種", "住所", "電話番号"])


# ===============================
# ★ 新プロファイル用のヘルパー（ヘッダーなし・業種＋住所同セル）
# ===============================
JP_LOC_PATTERN = re.compile(r"(丁目|番地?|号|市|区|町|村|郡|県|府|道)")

def is_hours_or_business_line(text: str) -> bool:
    """営業時間・診療時間系の行かどうか（住所候補からは除外）"""
    t = normalize_text(text)
    if not t:
        return False
    keywords = [
        "営業時間", "営業中", "営業時間外", "営業開始",
        "まもなく営業開始", "診療時間", "診察時間", "24時間営業",
    ]
    return any(k in t for k in keywords)

def is_address_like(text: str) -> bool:
    """住所らしいかどうかのゆるい判定（Google縦型の旧ロジック用）"""
    t = normalize_text(text)
    if not t:
        return False

    # ★ 営業時間系の行は住所扱いしない
    if is_hours_or_business_line(t):
        return False

    has_digit = bool(re.search(r"\d", t))
    has_loc_word = bool(JP_LOC_PATTERN.search(t))
    has_block = bool(re.search(r"\d{1,3}[-－ー‐]\d{1,3}", t))

    if has_digit and (has_loc_word or has_block):
        return True

    # 数字がなくても「○○市」「○○町」など住所語だけのケースを弱めに許可
    if has_loc_word and not has_digit:
        return True

    return False

def split_industry_address(text: str):
    """セル内の右端の「·/・/･」で業種と住所に分割"""
    t = normalize_text(text)
    if not t:
        return "", ""
    # 右から1つ目の区切りを探す
    last_pos = -1
    for ch in ["·", "・", "･"]:
        p = t.rfind(ch)
        if p > last_pos:
            last_pos = p
    if last_pos == -1:
        # 区切りがなければ全体を住所扱い
        return "", t.strip()
    left = t[:last_pos].strip()
    right = t[last_pos + 1 :].strip()
    if not right:
        # 右側が空なら住所扱いに倒す
        return "", left
    return left, right

KANJI_KATA_HIRA = r"\u4E00-\u9FFF\u30A0-\u30FF\u3040-\u309F"

def is_company_candidate(text: str) -> bool:
    """企業名として使えそうかどうか"""
    s = normalize_text(text)
    if not s:
        return False

    # 無視したいキーワード
    noise_words = [
        "ウェブサイト", "Web サイト", "web サイト",
        "オンラインで予約",
        "ルート・乗換", "経路案内",
        "共有",
        "営業中", "営業時間", "営業時間外", "営業開始",
        "まもなく営業開始", "クチコミはありません",
        "口コミ", "クチコミ", "レビュー", "件の",
    ]
    if any(w in s for w in noise_words):
        return False

    # レビュー点数形式: 5.0(1) など
    if re.match(r"^\d+(?:\.\d+)?\s*\(.+\)\s*$", s):
        return False

    # 数値や記号のみ (-22, 3.5 など) を除外
    if re.match(r"^[\d\.\-＋\+マイナス\s]+$", s):
        return False

    # ひらがな・カタカナ・漢字・英字が少なくとも1つ
    if not re.search(rf"[{KANJI_KATA_HIRA}A-Za-z]", s):
        return False

    return True

def is_google_meta_line(text: str) -> bool:
    """Google検索結果に出てくるメタ情報行かどうか（住所・業種候補からは除外）"""
    t = normalize_text(text)
    if not t:
        return True  # 空行はメタ扱いで飛ばす

    meta_keywords = [
        "ルート・乗換", "経路案内",
        "ウェブサイト", "Web サイト", "web サイト",
        "オンラインで予約",
        "共有",
        "現在営業中", "営業時間", "営業時間外",
        "営業開始", "まもなく営業開始", "24時間営業",
        "クチコミはありません", "口コミ", "クチコミ", "レビュー",
    ]
    if any(k in t for k in meta_keywords):
        return True

    # 数値や記号だけの行（評価点、-22 など）
    if re.match(r"^[\d\.\-＋\+マイナス\s]+$", t):
        return True

    return False

def extract_google_free_vertical(df_like: pd.DataFrame) -> pd.DataFrame:
    """
    Google検索結果（縦並び・ヘッダーなし・
    「業種＋住所」が同じセルに入っているパターン）から

      企業名 / 業種 / 住所 / 電話番号

    を抽出する。
    企業名は「電話から3〜4行上」のルールを優先しつつ、
    その間の行から業種＋住所のセルを拾う。
    """
    df0 = df_like.fillna("")
    col = df0.iloc[:, 0].astype(str).tolist()
    n = len(col)
    results = []

    for i, line in enumerate(col):
        ph_raw = pick_phone_token_raw(line)
        if not ph_raw:
            continue
        phone = ph_raw

        # --------------------------
        # 1) 企業名の行を決める
        # --------------------------
        company_idx = None

        # まず Jin さんルールで候補を決める
        txt_m2 = normalize_text(col[i - 2]) if i - 2 >= 0 else ""
        if i - 3 >= 0 and "クチコミはありません" in txt_m2:
            # 電話の2行上に「クチコミはありません」→ 3行上が企業名候補
            company_idx = i - 3
        elif i - 4 >= 0:
            # それ以外は基本4行上
            company_idx = i - 4

        # 候補が会社名として微妙なら、上方向にスキャンして会社名らしい行を探す
        if company_idx is not None:
            if not is_company_candidate(col[company_idx]):
                company_idx = None

        if company_idx is None:
            for k in range(i - 1, -1, -1):
                if is_company_candidate(col[k]):
                    company_idx = k
                    break

        if company_idx is None:
            # 企業名がどうしても見つからない場合はこの電話はスキップ
            continue

        company = normalize_text(col[company_idx])

        # --------------------------
        # 2) 業種＋住所セルを探す
        # --------------------------
        indaddr_idx = None
        # 電話の1行上から企業名の1行下までを逆順に見て、
        # メタ行を飛ばしながら最初に見つかった行を採用
        for j in range(i - 1, company_idx, -1):
            txt = normalize_text(col[j])
            if not txt:
                continue
            if is_google_meta_line(txt):
                continue
            indaddr_idx = j
            break

        # どうしても見つからない場合の保険として、
        # 電話の1行上から上方向にメタ以外の行を探す
        if indaddr_idx is None:
            for j in range(i - 1, -1, -1):
                txt = normalize_text(col[j])
                if not txt:
                    continue
                if is_google_meta_line(txt):
                    continue
                indaddr_idx = j
                break

        industry = ""
        address = ""

        if indaddr_idx is not None:
            ind_raw, addr_raw = split_industry_address(col[indaddr_idx])

            if addr_raw:
                # 「業種・住所」のように分割できたケース
                industry = extract_industry(ind_raw)
                address = clean_address(addr_raw)
            else:
                # 区切り記号が無い → 全体を住所扱い
                address = clean_address(col[indaddr_idx])

        # --------------------------
        # 3) 結果として追加
        # --------------------------
        results.append([company, industry, address, phone])

    if not results:
        return pd.DataFrame(columns=["企業名", "業種", "住所", "電話番号"])

    return pd.DataFrame(results, columns=["企業名", "業種", "住所", "電話番号"])


# ===============================
# 業種のフィルター/ハイライト
# ===============================
remove_exact = [
    "オフィス機器レンタル業", "足場レンタル会社", "電気工", "廃棄物リサイクル業",
    "プロパン販売業者", "看板専門店", "給水設備工場", "警備業", "建設会社",
    "工務店", "写真店", "人材派遣業", "整備店", "倉庫", "肉店", "米販売店",
    "スーパーマーケット", "ロジスティクスサービス", "建材店",
    "自動車整備工場", "自動車販売店", "車体整備店", "協会/組織", "建設請負業者", "電器店", "家電量販店", "建築会社", "ハウス クリーニング業", "焼肉店",
    "建築設計事務所", "左官", "作業服店", "空調設備工事業者", "金属スクラップ業者", "害獣駆除サービス", "モーター修理店", "アーチェリーショップ", "アスベスト検査業", "事務用品店",
    "測量士", "配管業者", "労働組合", "ガス会社", "ガソリンスタンド", "ガラス/ミラー店", "ワイナリー", "屋根ふき業者", "高等学校", "金物店", "史跡", "商工会議所", "清掃業", "清掃業者", "配管工", "お手頃"
]
remove_partial = ["販売店", "販売業者"]

highlight_partial = [
    "運輸", "ロジスティクスサービス", "倉庫", "輸送サービス",
    "運送会社企業のオフィス", "運送会社"
]

# ===============================
# 業種ノイズ除去（レビュー/評価など）
# ===============================
def clean_industry_noise(s: str) -> str:
    """
    業種カラムに紛れ込む
    - レビュー情報（レビュー・なし・…）
    - Google のクチコミ
    - ○件のレビュー／口コミ
    などのノイズを除去する
    ＋ 最後に「·」「レビュ-なし」「空白だけ」は必ず消す
    """
    if not s:
        return ""
    t = str(s)
    # 空白をゆるく正規化
    t = re.sub(r"\s+", " ", t).strip()

    # 先頭の評価スコア + 件数 例: '4.7(123)・', '4.7（123）・'
    t = re.sub(r"^\s*\d+(?:\.\d+)?\s*[\(（]\s*\d+\s*[\)）]\s*(?:件)?\s*[・･]?\s*", "", t)

    # ---- 「レビュー・なし・○○」系をトークン単位で処理 ----
    def norm_token(x: str) -> str:
        return re.sub(r"\s+", "", x)

    noise_basic = {"レビュー", "レビューなし", "レビュー無し", "クチコミ", "口コミ"}
    noise_nashi = {"なし"}

    if t.startswith("レビュー"):
        parts = [p.strip() for p in re.split(r"[・･]", t) if p.strip()]
        if not parts:
            return ""

        # 全部ノイズなら空にする
        if all(norm_token(p) in noise_basic | noise_nashi for p in parts):
            return ""

        cleaned_parts = []
        for p in parts:
            pn = norm_token(p)
            if pn in noise_basic or pn in noise_nashi:
                continue
            cleaned_parts.append(p)

        t = "・".join(cleaned_parts)
    else:
        # 「Google のクチコミ」「口コミ」「クチコミ」などが途中にある場合
        t = re.sub(r"(?:^|[・･])\s*(Google\s*の?\s*クチコミ|口コミ|クチコミ)\s*(?=[・･]|$)", "", t)
        # 「◯件のレビュー」「◯件の口コミ」など
        t = re.sub(r"[・･]?\s*\d+\s*件の?(レビュー|口コミ|クチコミ)\s*(?=[・･]|$)", "", t)

    # 分割して空要素を削除
    parts = [p.strip() for p in re.split(r"[・･]", t) if p.strip()]
    t = "・".join(parts) if parts else ""

    # 余計な区切りや空白を整形
    t = re.sub(r"[・･]{2,}", "・", t).strip(" ・･")

    # ▼▼▼ ここが「必ず消す」部分 ▼▼▼
    # 中黒「·」や「レビュ-なし」を強制削除
    if t:
        for trash in ["·", "レビュ-なし"]:
            t = t.replace(trash, "")
        # ついでに全角/半角スペースだけになった場合も空にする
        t = re.sub(r"\s+", " ", t).strip()

    return t if t else ""

# ===============================
# 共通整形（電話は触らない）
# ===============================
def clean_dataframe_except_phone(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in ["企業名", "業種", "住所"]:
        df[c] = df[c].map(normalize_text)
    df["業種"] = df["業種"].map(clean_industry_noise)
    return df.fillna("")

# ===============================
# UI（NGリスト選択・抽出方式・業種カテゴリ・テンプレート入力）
# ===============================
st.markdown("### 🛡️ 使用するNGリストを選択")
nglist_files = [f for f in os.listdir() if f.endswith(".xlsx") and "NGリスト" in f]
nglist_options = ["なし"] + [os.path.splitext(f)[0] for f in nglist_files]
selected_nglist = st.selectbox(
    "NGリスト",
    nglist_options,
    index=0,
    help="同じフォルダにある『NGリスト〜.xlsx』を検出します。1列目=企業名、2列目=電話番号（任意）。"
)

st.markdown("### 🧭 抽出方法を選択")
profile = st.selectbox(
    "抽出プロファイル",
    [
        "Google検索リスト（縦読み・電話上下型）",
        "Google検索リスト（ヘッダーなし・業種＋住所同セル）",  # ★追加
        "シゴトアルワ検索リスト（縦積み）",
        "日本倉庫協会リスト（4列型）",
    ]
)

st.markdown("### 🏭 業種カテゴリを選択")
industry_option = st.radio("どの業種カテゴリーに該当しますか？", ("製造業", "物流業", "その他"))

st.markdown("### 🧩 テンプレートの取得方法（OS互換強化）")
template_source = st.radio(
    "template.xlsx の取得元",
    ("プロジェクト内の template.xlsx を使う（従来）", "ここで template.xlsx をアップロードして使う"),
    index=0
)
template_upload = None
if template_source == "ここで template.xlsx をアップロードして使う":
    template_upload = st.file_uploader("template.xlsx をアップロード", type=["xlsx"], key="template_up")

# ★ 複数ファイル対応：accept_multiple_files=True
uploaded_files = st.file_uploader(
    "📤 整形対象のExcelファイルをアップロード（複数選択可）",
    type=["xlsx"],
    accept_multiple_files=True
)

# ===============================
# NGリストを一度だけ読み込んで共有
# ===============================
ng_names = []
ng_phones = set()
if uploaded_files and selected_nglist != "なし":
    ng_path = f"{selected_nglist}.xlsx"
    if not os.path.exists(ng_path):
        st.error(f"❌ 選択されたNGリストが見つかりません：{ng_path}")
        st.stop()
    ng_df = pd.read_excel(ng_path, engine="openpyxl").fillna("")
    if ng_df.shape[1] < 1:
        st.error("❌ NGリストは少なくとも1列（企業名）が必要です。2列目に電話番号があれば照合に利用します。")
        st.stop()
    ng_df["__ng_company_canon"] = ng_df.iloc[:, 0].map(canonical_company_name)
    if ng_df.shape[1] >= 2:
        ng_df["__ng_digits"] = ng_df.iloc[:, 1].astype(str).map(phone_digits_only)
    else:
        ng_df["__ng_digits"] = ""
    ng_names = [n for n in ng_df["__ng_company_canon"].tolist() if n]
    ng_phones = set([d for d in ng_df["__ng_digits"].tolist() if d])

# ===============================
# メイン処理（★ファイルごとに独立して処理）
# ===============================
if uploaded_files:
    for file_index, uploaded_file in enumerate(uploaded_files):
        st.markdown("---")
        st.markdown(f"## 📁 {uploaded_file.name}")

        filename_no_ext = os.path.splitext(uploaded_file.name)[0]
        xl = pd.ExcelFile(uploaded_file, engine="openpyxl")

        # --- 抽出 ---
        if "入力マスター" in xl.sheet_names:
            # template互換: 入力マスターから読み取り（電話は原文のまま）
            df_raw = pd.read_excel(
                xl,
                sheet_name="入力マスター",
                header=None,
                engine="openpyxl"
            ).fillna("")
            df = pd.DataFrame({
                "企業名": df_raw.iloc[1:, 1].astype(str),
                "業種": df_raw.iloc[1:, 2].astype(str),
                "住所": df_raw.iloc[1:, 3].astype(str),
                "電話番号": df_raw.iloc[1:, 4].astype(str),
            })
        else:
            df0 = pd.read_excel(uploaded_file, header=None, engine="openpyxl").fillna("")
            if profile == "Google検索リスト（縦読み・電話上下型）":
                lines = df0.iloc[:, 0].tolist()
                df = extract_google_vertical(lines)
            elif profile == "Google検索リスト（ヘッダーなし・業種＋住所同セル）":
                df = extract_google_free_vertical(df0)
            elif profile == "シゴトアルワ検索リスト（縦積み）":
                df = extract_shigoto_arua(df0)
            else:
                df = extract_warehouse_association(df0)

        # --- 非電話列のみ正規化 ---
        df = clean_dataframe_except_phone(df)

        # --- 比較キー ---
        df["__company_canon"] = df["企業名"].map(canonical_company_name)
        df["__digits"] = df["電話番号"].map(phone_digits_only)

        # --- 業種フィルター（製造業のみ除外ルール適用） ---
        removed_by_industry = 0
        if industry_option == "製造業":
            before = len(df)
            all_ng_words = remove_exact + remove_partial
            if all_ng_words:
                pat = "|".join(map(re.escape, all_ng_words))
                df = df[~df["業種"].str.contains(pat, na=False)]
            removed_by_industry = before - len(df)
            st.warning(f"🏭 製造業フィルター適用：{removed_by_industry}件を除外しました")

        # --- NG照合（任意） ---
        removal_logs = []
        company_removed = 0
        phone_removed = 0
        dup_removed = 0

        if ng_names or ng_phones:
            # 企業名（部分一致・相互包含）
            before = len(df)
            hit_idx = []
            for idx, row in df.iterrows():
                c = row["__company_canon"]
                if not c:
                    continue
                if any((n in c or c in n) for n in ng_names):
                    removal_logs.append({
                        "reason": "ng-company",
                        "company": row["企業名"],
                        "phone_raw": row["電話番号"],
                        "match": c
                    })
                    hit_idx.append(idx)
            if hit_idx:
                df = df.drop(index=hit_idx)
            company_removed = before - len(df)

            # 電話番号digits一致
            before = len(df)
            mask = df["__digits"].isin(ng_phones)
            if mask.any():
                for idx, row in df[mask].iterrows():
                    removal_logs.append({
                        "reason": "ng-phone",
                        "company": row["企業名"],
                        "phone_raw": row["電話番号"],
                        "match": row["__digits"]
                    })
                df = df[~mask]
            phone_removed = before - len(df)

        # --- 重複（電話digits）除去（※このファイル内だけ） ---
        before = len(df)
        dup_mask = df["__digits"].ne("").astype(bool) & df["__digits"].duplicated(keep="first")
        if dup_mask.any():
            for idx, row in df[dup_mask].iterrows():
                removal_logs.append({
                    "reason": "dup-phone",
                    "company": row["企業名"],
                    "phone_raw": row["電話番号"],
                    "match": row["__digits"]
                })
            df = df[~dup_mask]
        dup_removed = before - len(df)

        # --- 空行の除去 ---
        df = df[~((df["企業名"] == "") & (df["業種"] == "") & (df["住所"] == "") & (df["電話番号"] == ""))].reset_index(drop=True)

        # --- 画面表示（編集可・確定ボタンなし） ---
        st.success(f"✅ 整形完了：{len(df)}件の企業データを取得しました。")
        edited = st.data_editor(
            df[["企業名", "業種", "住所", "電話番号"]],
            use_container_width=True,
            num_rows="fixed",
            column_config={
                "企業名": st.column_config.TextColumn(required=True),
                "業種": st.column_config.TextColumn(),
                "住所": st.column_config.TextColumn(),
                "電話番号": st.column_config.TextColumn(
                    help="原文の配列を保持。必要ならここで手動修正してください。編集内容はそのまま出力に反映されます。"
                ),
            },
            key=f"editable_preview_{file_index}",
        )

        # 確定ボタンは廃止。edited をそのまま出力用に使う
        df_export = edited.copy()

        # --- サマリー＆削除ログDL ---
        with st.expander(f"📊 実行サマリー（詳細） - {uploaded_file.name}", expanded=False):
            st.markdown(
                f"- フィルター除外（製造業 部分一致）: **{removed_by_industry}** 件\n"
                f"- NG（企業名 部分一致）削除: **{company_removed}** 件\n"
                f"- NG（電話 digits一致）削除: **{phone_removed}** 件\n"
                f"- 重複（電話 digits一致）削除: **{dup_removed}** 件\n"
            )
            if removal_logs:
                log_df = pd.DataFrame(removal_logs)
                st.dataframe(log_df.head(300), use_container_width=True)
                csv_bytes = log_df.to_csv(index=False).encode("utf-8-sig")
                st.download_button(
                    "🧾 削除ログをCSVでダウンロード",
                    data=csv_bytes,
                    file_name=f"removal_logs_{filename_no_ext}.csv",
                    mime="text/csv",
                    key=f"removal_log_btn_{file_index}",
                )

        # ===============================
        # template.xlsx へ書き込み（OS互換強化）
        # ===============================
        wb = None
        if template_upload is not None:
            try:
                buf = io.BytesIO(template_upload.read())
                wb = load_workbook(buf)
            except Exception as e:
                st.error(f"❌ アップロードした template.xlsx の読み込みに失敗しました: {e}")
                st.stop()
        else:
            app_dir = Path(__file__).resolve().parent
            template_path = app_dir / "template.xlsx"
            if not template_path.exists():
                st.error(
                    f"❌ template.xlsx が見つかりませんでした（期待パス: {template_path}）。"
                    "『ここで template.xlsx をアップロードして使う』を選ぶか、"
                    "ファイルをプロジェクト直下に配置してください。"
                )
                st.stop()
            try:
                wb = load_workbook(template_path)
            except Exception as e:
                st.error(f"❌ template.xlsx の読み込みに失敗しました: {e}")
                st.stop()

        if "入力マスター" not in wb.sheetnames:
            st.error("❌ template.xlsx に『入力マスター』というシートが存在しません。")
            st.stop()

        sheet_master = wb["入力マスター"]

        # 既存データ（2行目以降のB〜E）と塗りをクリア
        for row in sheet_master.iter_rows(min_row=2, max_row=sheet_master.max_row):
            for cell in row[1:5]:  # B(1)〜E(4)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)

        # 物流ハイライト（業種に特定語が含まれる場合、C列を赤く）
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        def is_logi(val: str) -> bool:
            v = (val or "").strip()
            return any(word in v for word in highlight_partial)

        # データ書き込み（B=企業名, C=業種, D=住所, E=電話）
        for idx_row, row in df_export.iterrows():
            r = idx_row + 2
            sheet_master.cell(row=r, column=2, value=row["企業名"])
            sheet_master.cell(row=r, column=3, value=row["業種"])
            sheet_master.cell(row=r, column=4, value=row["住所"])
            sheet_master.cell(row=r, column=5, value=row["電話番号"])
            if industry_option == "物流業" and is_logi(row["業種"]):
                sheet_master.cell(row=r, column=3).fill = red_fill

        # ===============================
        # 開拓先リストシートのプルダウン＆印刷範囲設定
        # ===============================
        if "開拓先リスト" in wb.sheetnames:
            sheet_k = wb["開拓先リスト"]

            # プルダウン（データ検証）: H列の H3, H9, H15, ... に設定
            try:
                dv = DataValidation(
                    type="list",
                    formula1='"-,アポ,見込み,断り,留守,担当者不在,不使用,削除依頼"',
                    allow_blank=True,
                )
                sheet_k.add_data_validation(dv)

                max_row_k = sheet_k.max_row or 200
                row = 3
                while row <= max_row_k:
                    cell_ref = f"H{row}"
                    dv.add(sheet_k[cell_ref])
                    row += 6
            except Exception:
                # DataValidation がうまく行かない場合は何もしない（エラーで止めない）
                pass

            # 印刷範囲を A〜L 全行に設定
            try:
                max_row_k = sheet_k.max_row or 200
                sheet_k.print_area = f"A1:L{max_row_k}"
            except Exception:
                pass

        # ダウンロード（ファイルごとに別ボタン）
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button(
            label=f"📥 整形済みリストをダウンロード（{filename_no_ext} / template.xlsx 反映）",
            data=output,
            file_name=f"{filename_no_ext}リスト.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_btn_{file_index}",
        )

else:
    st.info("Excelファイルをアップロードしてください。NGリストxlsxは同フォルダに置くか、プロジェクト直下に配置してください。")
